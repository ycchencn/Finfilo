"""
 * @author Yc
 * Chaos isn't a pit. Chaos is a ladder. - Littlefinger
 * Copyright (c) 2025 yccheni@163.com. All rights reserved.
"""

import pandas as pd
import time
from openpyxl.styles import Alignment
from utils.data_loader import datagigi
from service import ResearchReportService
from service.stock import StockService


def export_dcf_to_excel(
        output_path='dcf_valuation_report.xlsx',
        sort_by='中性空间',
        ascending=False
):
    """从监控股票池获取DCF估值信息，输出为Excel文件（纯数值，百分比列设置Excel格式，首行筛选，左对齐垂直居中）"""
    stocks = StockService.get_monitoring_stock_pool(market='cn', per_page=300)

    records = []

    for stock in stocks:
        time.sleep(0.1)
        try:
            symbol = stock['symbol']
            report = ResearchReportService.get_by_code(stock_code=symbol, report_type=1)
            if not report:
                print(f"[跳过] {symbol} 无DCF报告")
                continue

            content = report.get('content_json')
            if not content:
                print(f"[跳过] {symbol} 报告内容为空")
                continue

            valuation = content.get('每股内在价值')
            if not valuation:
                print(f"[跳过] {symbol} 无每股内在价值数据")
                continue

            try:
                opt = float(valuation.get('乐观情景', 0))
                mid = float(valuation.get('中性情景', 0))
                cons = float(valuation.get('保守情景', 0))
            except (ValueError, TypeError) as e:
                print(f"[错误] {symbol} 估值数据格式异常: {e}")
                continue

            last_tick = datagigi.get_last_tick(symbol=symbol)

            if last_tick and 'lastPrice' in last_tick:
                current_price = float(last_tick['lastPrice'])
            else:
                price_str = content.get('当前股价')
                if price_str:
                    current_price = float(price_str)
                else:
                    print(f"[跳过] {symbol} 无法获取现价")
                    continue

            stock_info = datagigi.get_stock_info(symbol=symbol)
            name = stock_info.get('name', 'N/A') if stock_info else 'N/A'

            if current_price == 0:
                print(f"[跳过] {symbol} 现价为0")
                continue

            # 存储为小数（例如0.377代表37.7%）
            opt_space = (opt - current_price) / current_price
            mid_space = (mid - current_price) / current_price
            cons_space = (cons - current_price) / current_price

            records.append({
                '名称': name,
                '代码': symbol,
                '现价': current_price,
                '乐观估值': opt,
                '乐观空间': opt_space,
                '中性估值': mid,
                '中性空间': mid_space,
                '保守估值': cons,
                '保守空间': cons_space
            })

            print(f"处理数据，{symbol}, {name}")

        except Exception as e:
            print(f"[异常] 处理 {stock.get('symbol', '未知')} 时出错: {e}")
            continue

    if not records:
        print("没有有效的记录，无法生成Excel文件")
        return

    df = pd.DataFrame(records)

    if sort_by in df.columns:
        df = df.sort_values(by=sort_by, ascending=ascending)
    else:
        print(f"[警告] 排序列 '{sort_by}' 不存在，将按原始顺序输出")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='DCF估值报告')
        worksheet = writer.sheets['DCF估值报告']

        # 1. 设置百分比列的单元格格式
        percent_columns = ['乐观空间', '中性空间', '保守空间']
        for idx, col_name in enumerate(df.columns, start=1):
            if col_name in percent_columns:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=idx)
                    cell.number_format = '0.00%'

        # 2. 设置所有单元格左对齐、垂直居中
        for row in range(1, worksheet.max_row + 1):           # 包含表头
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # 3. 设置列宽（固定15）
        for column_cells in worksheet.columns:
            worksheet.column_dimensions[column_cells[0].column_letter].width = 15

        # 4. 开启首行自动筛选
        worksheet.auto_filter.ref = worksheet.dimensions

    print(f"✅ 报告已生成：{output_path}")


if __name__ == '__main__':
    export_dcf_to_excel('dcf_valuation_report.xlsx', sort_by='中性空间', ascending=False)